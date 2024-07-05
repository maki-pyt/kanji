#!/usr/bin/env python
# coding: utf-8

import openpyxl
import streamlit as st
import random
import os


st.header('漢検対策プリント', divider='blue')

page = st.sidebar.selectbox('漢字プリントを作成する。', ['7級', '6級', '5級'], index=0)
 
if page == '7級':
    st.title('7級用のプリント作成ページです。')

    wb = openpyxl.load_workbook("kanjimondai2.xlsx")
    ws = wb["7級"]
    # シート名で指定する場合
    # １行目（列名のセル）
    header_cells = ws[1]

    # ２行目以降（データ）
    question_list = []
    for row in ws.iter_rows(min_row=2):
        row_dic = {}
    # セルの値を「key-value」で登録
        for k, v in zip(header_cells, row):
            row_dic[k.value] = v.value
        question_list.append(row_dic)
    max = len(question_list)
    col1, col2 = st.columns(2)  # 2列のコンテナを用意する
    with col1:
          start = st.number_input('ここから', min_value=1, max_value=max, value=1)

    with col2:
          finish = st.number_input('ここまで', min_value=20, max_value=max, value=20)

    a = start
    b = finish+2
    pre_test_list = question_list[a:b]
    test_list = random.sample(pre_test_list, 20)

    clm1 = [s["前文"] for s in test_list]
    clm2 = [s["漢字（回答部分）"] for s in test_list]
    clm3 = [s["ひらがな（回答部分）"] for s in test_list]
    clm4 = [s["後文"] for s in test_list]
    
    ws = wb['7級用プリント作成シート']

    for i in range(0,len(clm1)):
    #列に書き込み
        ws.cell(i+1,1,value = clm1[i])
        ws.cell(i+1,2,value = clm2[i])
        ws.cell(i+1,3,value = clm3[i])
        ws.cell(i+1,4,value = clm4[i])

    wb = openpyxl.load_workbook("kanjiprint2.xlsx")
    ws = wb['プリント作成用シート']

    for i in range(0,len(clm1)):
    #列に書き込み
        ws.cell(i+1,1,value = clm1[i])
        ws.cell(i+1,2,value = clm2[i])
        ws.cell(i+1,3,value = clm3[i])
        ws.cell(i+1,4,value = clm4[i])

    sheet3=wb['プリント作成用シート']
    #Sheet1の値のある行数を取得
    rw=sheet3.max_row
    #Sheet1の値のある列数を取得
    cl=sheet3.max_column

    sheet4=wb['プリント作成用シート2']
    #iは値のある行数分繰り返す
    #jは値のある列数分繰り返す
    #range(start,stop)はstart≦i<stopでstopで指定した値は含まないので「+1」している
    for i in range(1,rw+1):
        for j in range(1,cl+1):
            C1=sheet3.cell(row=i,column=j) #sheet1のセルの行番号と列番号を指定している
            C2=sheet4.cell(row=j,column=i) #sheet1のセルの行番号と列番号を入れ替えてsheet2のセルを指定している
            C2.value=C1.value #sheet2のセルにsheet1のセルの値を代入

    wb.save('kanjiprint2.xlsx') #上書き保存
    if st.button('プリント作成'):
       st.write('完成しました。以下のURLから印刷して使用してください。オリジナルの漢字プリントを作りたい場合は下のファイルをダウンロードして、コピーを保存しプリント作成用シートにデータを貼り付けて作成してください。')
       st.markdown('https://docs.google.com/spreadsheets/d/1qv4GptxvTXnqGNWdr4MPwoyhQ5PmFh_W/edit?usp=share_link&ouid=104200975424459618460&rtpof=true&sd=true',unsafe_allow_html=True)

elif page == '6級':
    st.title('6級用のプリント作成ページです。')

    wb = openpyxl.load_workbook("kanjimondai2.xlsx")
    ws = wb["7級"]
    # シート名で指定する場合
    # １行目（列名のセル）
    header_cells = ws[1]

    # ２行目以降（データ）
    question_list = []
    for row in ws.iter_rows(min_row=2):
        row_dic = {}
    # セルの値を「key-value」で登録
        for k, v in zip(header_cells, row):
            row_dic[k.value] = v.value
        question_list.append(row_dic)
    max = len(question_list)
    col1, col2 = st.columns(2)  # 2列のコンテナを用意する
    with col1:
          start = st.number_input('ここから', min_value=1, max_value=max, value=1)

    with col2:
          finish = st.number_input('ここまで', min_value=20, max_value=max, value=20)

    a = start
    b = finish+2
    pre_test_list = question_list[a:b]
    test_list = random.sample(pre_test_list, 20)

    clm1 = [s["前文"] for s in test_list]
    clm2 = [s["漢字（回答部分）"] for s in test_list]
    clm3 = [s["ひらがな（回答部分）"] for s in test_list]
    clm4 = [s["後文"] for s in test_list]
    
    ws = wb['6級用プリント作成シート']

    for i in range(0,len(clm1)):
    #列に書き込み
        ws.cell(i+1,1,value = clm1[i])
        ws.cell(i+1,2,value = clm2[i])
        ws.cell(i+1,3,value = clm3[i])
        ws.cell(i+1,4,value = clm4[i])

    wb = openpyxl.load_workbook("kanjiprint2.xlsx")
    ws = wb['プリント作成用シート']

    for i in range(0,len(clm1)):
    #列に書き込み
        ws.cell(i+1,1,value = clm1[i])
        ws.cell(i+1,2,value = clm2[i])
        ws.cell(i+1,3,value = clm3[i])
        ws.cell(i+1,4,value = clm4[i])

    sheet3=wb['プリント作成用シート']
    #Sheet1の値のある行数を取得
    rw=sheet3.max_row
    #Sheet1の値のある列数を取得
    cl=sheet3.max_column

    sheet4=wb['プリント作成用シート2']
    #iは値のある行数分繰り返す
    #jは値のある列数分繰り返す
    #range(start,stop)はstart≦i<stopでstopで指定した値は含まないので「+1」している
    for i in range(1,rw+1):
        for j in range(1,cl+1):
            C1=sheet3.cell(row=i,column=j) #sheet1のセルの行番号と列番号を指定している
            C2=sheet4.cell(row=j,column=i) #sheet1のセルの行番号と列番号を入れ替えてsheet2のセルを指定している
            C2.value=C1.value #sheet2のセルにsheet1のセルの値を代入

    wb.save('kanjiprint2.xlsx') #上書き保存
    if st.button('プリント作成'):
       st.write('完成しました。以下のURLから印刷して使用してください。オリジナルの漢字プリントを作りたい場合は下のファイルをダウンロードして、コピーを保存しプリント作成用シートにデータを貼り付けて作成してください。')
       st.markdown('https://docs.google.com/spreadsheets/d/1qv4GptxvTXnqGNWdr4MPwoyhQ5PmFh_W/edit?usp=share_link&ouid=104200975424459618460&rtpof=true&sd=true',unsafe_allow_html=True)

elif page == '5級':
    st.title('5級用のプリント作成ページです。')

    wb = openpyxl.load_workbook("kanjimondai2.xlsx")
    ws = wb["5級"]
    # シート名で指定する場合
    # １行目（列名のセル）
    header_cells = ws[1]

    # ２行目以降（データ）
    question_list = []
    for row in ws.iter_rows(min_row=2):
        row_dic = {}
    # セルの値を「key-value」で登録
        for k, v in zip(header_cells, row):
            row_dic[k.value] = v.value
        question_list.append(row_dic)
    max = len(question_list)
    col1, col2 = st.columns(2)  # 2列のコンテナを用意する
    with col1:
          start = st.number_input('ここから', min_value=1, max_value=max, value=1)

    with col2:
          finish = st.number_input('ここまで', min_value=20, max_value=max, value=20)

    a = start
    b = finish+2
    pre_test_list = question_list[a:b]
    test_list = random.sample(pre_test_list, 20)

    clm1 = [s["前文"] for s in test_list]
    clm2 = [s["漢字（回答部分）"] for s in test_list]
    clm3 = [s["ひらがな（回答部分）"] for s in test_list]
    clm4 = [s["後文"] for s in test_list]
    
    ws = wb['5級用プリント作成シート']

    for i in range(0,len(clm1)):
    #列に書き込み
        ws.cell(i+1,1,value = clm1[i])
        ws.cell(i+1,2,value = clm2[i])
        ws.cell(i+1,3,value = clm3[i])
        ws.cell(i+1,4,value = clm4[i])

    wb = openpyxl.load_workbook("kanjiprint2.xlsx")
    ws = wb['プリント作成用シート']

    for i in range(0,len(clm1)):
    #列に書き込み
        ws.cell(i+1,1,value = clm1[i])
        ws.cell(i+1,2,value = clm2[i])
        ws.cell(i+1,3,value = clm3[i])
        ws.cell(i+1,4,value = clm4[i])

    sheet3=wb['プリント作成用シート']
    #Sheet1の値のある行数を取得
    rw=sheet3.max_row
    #Sheet1の値のある列数を取得
    cl=sheet3.max_column

    sheet4=wb['プリント作成用シート2']
    #iは値のある行数分繰り返す
    #jは値のある列数分繰り返す
    #range(start,stop)はstart≦i<stopでstopで指定した値は含まないので「+1」している
    for i in range(1,rw+1):
        for j in range(1,cl+1):
            C1=sheet3.cell(row=i,column=j) #sheet1のセルの行番号と列番号を指定している
            C2=sheet4.cell(row=j,column=i) #sheet1のセルの行番号と列番号を入れ替えてsheet2のセルを指定している
            C2.value=C1.value #sheet2のセルにsheet1のセルの値を代入

    wb.save('kanjiprint2.xlsx') #上書き保存
    if st.button('プリント作成'):
       st.write('完成しました。以下のURLから印刷して使用してください。オリジナルの漢字プリントを作りたい場合は下のファイルをダウンロードして、コピーを保存しプリント作成用シートにデータを貼り付けて作成してください。')
       st.markdown('https://docs.google.com/spreadsheets/d/1qv4GptxvTXnqGNWdr4MPwoyhQ5PmFh_W/edit?usp=share_link&ouid=104200975424459618460&rtpof=true&sd=true',unsafe_allow_html=True)

else:
    st.write('作成したい級を選んでください')
    
if st.button('問題リストを作成する。'):
   st.write('こちらの Googleスプレッドシートを編集し問題リストを作成してください。')
   st.markdown('https://docs.google.com/spreadsheets/d/12GOjNRojL-F_U8XMF4G1HXEPKmeYCrBZ/edit?gid=1421499505#gid=1421499505&rtpof=true&sd=true',unsafe_allow_html=True)
